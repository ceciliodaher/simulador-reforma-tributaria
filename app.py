import streamlit as st
import pandas as pd
import numpy as np
import json
import os
import base64
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.utils.dataframe import dataframe_to_rows
import base64
from io import BytesIO
from datetime import datetime
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.units import inch
from config import ConfiguracaoTributaria
from calculadoras import CalculadoraTributosAtuais, CalculadoraIVADual
from utils import (formatar_br, criar_grafico_comparativo, criar_grafico_aliquotas,
                  criar_grafico_transicao, criar_grafico_incentivos)


# Configuração da página
st.set_page_config(
    page_title="Simulador da Reforma Tributária",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)


# Função para inicializar a sessão
def inicializar_sessao():
    if 'config' not in st.session_state:
        st.session_state.config = ConfiguracaoTributaria()

    if 'calculadora_iva' not in st.session_state:
        st.session_state.calculadora_iva = CalculadoraIVADual(st.session_state.config)

    if 'resultados' not in st.session_state:
        st.session_state.resultados = {}

    if 'aliquotas_equivalentes' not in st.session_state:
        st.session_state.aliquotas_equivalentes = {}

    if 'memoria_calculo' not in st.session_state:
        st.session_state.memoria_calculo = {}

    # Dados do formulário (para persistência e exportação)
    if 'faturamento' not in st.session_state:
        st.session_state.faturamento = 0
    if 'custos_tributaveis' not in st.session_state:
        st.session_state.custos_tributaveis = 0
    if 'custos_simples' not in st.session_state:
        st.session_state.custos_simples = 0
    if 'creditos_anteriores' not in st.session_state:
        st.session_state.creditos_anteriores = 0
    if 'setor' not in st.session_state:
        st.session_state.setor = "padrao"
    if 'regime' not in st.session_state:
        st.session_state.regime = "real"
    if 'carga_atual' not in st.session_state:
        st.session_state.carga_atual = 25
    if 'aliquota_entrada' not in st.session_state:
        st.session_state.aliquota_entrada = 19
    if 'aliquota_saida' not in st.session_state:
        st.session_state.aliquota_saida = 19


# Função para adicionar incentivo
def adicionar_incentivo(tipo, dados):
    """Adiciona um novo incentivo fiscal à configuração."""
    descricao = dados.get("descricao", f"Incentivo {tipo.capitalize()}")
    tipo_incentivo = dados.get("tipo", "Nenhum")
    percentual = dados.get("percentual", 0) / 100
    perc_operacoes = dados.get("perc_operacoes", 100) / 100

    if tipo_incentivo != "Nenhum" and percentual <= 0:
        st.error("O percentual do incentivo deve ser maior que zero.")
        return False

    # Validar total de operações incentivadas (não aplicável para incentivos de apuração)
    if tipo != "apuracao":
        incentivos = st.session_state.config.icms_config[f"incentivos_{tipo}"]
        total_percentual = perc_operacoes

        for inc in incentivos:
            total_percentual += inc.get("percentual_operacoes", 0)

        if total_percentual > 1:
            st.error(f"O total de operações incentivadas ({formatar_br(total_percentual * 100)}%) excede 100%. "
                     "Ajuste os percentuais para que a soma não ultrapasse 100%.")
            return False

    # Criar objeto de incentivo
    incentivo = {
        "descricao": descricao,
        "tipo": tipo_incentivo,
        "percentual": percentual,
        "percentual_operacoes": perc_operacoes,
        "aplicavel_saidas": tipo == "saida",
        "aplicavel_entradas": tipo == "entrada",
        "aplicavel_apuracao": tipo == "apuracao"
    }

    # Adicionar à configuração
    st.session_state.config.icms_config[f"incentivos_{tipo}"].append(incentivo)
    return True


# Função para remover incentivo
def remover_incentivo(tipo, indice):
    """Remove um incentivo fiscal da configuração."""
    if 0 <= indice < len(st.session_state.config.icms_config[f"incentivos_{tipo}"]):
        st.session_state.config.icms_config[f"incentivos_{tipo}"].pop(indice)
        return True
    return False


# Função para exportar resultados
def exportar_resultados(formato):
    """Exporta os resultados para o formato especificado."""
    if not st.session_state.resultados:
        st.error("Execute uma simulação antes de exportar os resultados.")
        return None

    if formato == "excel":
        output = BytesIO()

        def exportar_excel():
            """Exporta os resultados para um arquivo Excel detalhado."""
            if not st.session_state.resultados:
                st.warning("Execute uma simulação antes de exportar os resultados.")
                return None

            try:
                # Criar um buffer em memória para o arquivo Excel
                output = io.BytesIO()

                # Criar workbook
                wb = Workbook()

                # Aba de Parâmetros
                ws_parametros = wb.active
                ws_parametros.title = "Parâmetros"

                # Adicionar título
                ws_parametros['A1'] = "Simulador da Reforma Tributária - IVA Dual (CBS/IBS)"
                ws_parametros['A1'].font = Font(bold=True, size=14)
                ws_parametros.merge_cells('A1:B1')

                # Data do relatório
                import datetime
                ws_parametros['A2'] = "Data do relatório:"
                ws_parametros['B2'] = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")

                # Cabeçalho dos parâmetros
                ws_parametros['A4'] = "Parâmetro"
                ws_parametros['B4'] = "Valor"

                # Estilo do cabeçalho
                for cell in ws_parametros['A4:B4'][0]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Obter dados da sessão
                # Aqui vamos assumir que os dados foram salvos no estado da sessão após a simulação
                faturamento = st.session_state.get('faturamento', 0)
                custos = st.session_state.get('custos_tributaveis', 0)
                custos_simples = st.session_state.get('custos_simples', 0)
                creditos_anteriores = st.session_state.get('creditos_anteriores', 0)
                setor = st.session_state.get('setor', 'padrao')
                regime = st.session_state.get('regime', 'real')
                carga_atual = st.session_state.get('carga_atual', 25)

                # Preencher dados dos parâmetros
                parametros = [
                    ["Faturamento Anual", f"R$ {formatar_br(faturamento)}"],
                    ["Custos Tributáveis", f"R$ {formatar_br(custos)}"],
                    ["Fornecedores do Simples", f"R$ {formatar_br(custos_simples)}"],
                    ["Créditos Anteriores", f"R$ {formatar_br(creditos_anteriores)}"],
                    ["Setor de Atividade", setor],
                    ["Regime Tributário", regime],
                    ["Carga Tributária Atual", f"{formatar_br(carga_atual)}%"],
                ]

                # Adicionar parâmetros
                for i, (param, valor) in enumerate(parametros, 5):
                    ws_parametros[f'A{i}'] = param
                    ws_parametros[f'B{i}'] = valor

                # Ajustar largura das colunas
                ws_parametros.column_dimensions['A'].width = 25
                ws_parametros.column_dimensions['B'].width = 25

                # Aba de Resultados
                ws_resultados = wb.create_sheet(title="Resultados")

                # Adicionar título
                ws_resultados['A1'] = "Resultados da Simulação"
                ws_resultados['A1'].font = Font(bold=True, size=14)
                ws_resultados.merge_cells('A1:I1')

                # Cabeçalho dos resultados
                cabecalhos = [
                    "Ano", "CBS (R$)", "IBS (R$)", "Imposto Bruto (R$)", "Créditos (R$)",
                    "Imposto Devido (R$)", "Carga Atual (R$)", "Diferença (R$)",
                    "Alíquota Efetiva (%)"
                ]

                for col, header in enumerate(cabecalhos, 1):
                    cell = ws_resultados.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Preparar dados
                dados_tabela = []
                for ano, resultado in st.session_state.resultados.items():
                    valor_atual = st.session_state.aliquotas_equivalentes[ano]["valor_atual"]
                    diferenca = resultado["imposto_devido"] - valor_atual

                    dados_tabela.append({
                        "Ano": ano,
                        "CBS (R$)": resultado["cbs"],
                        "IBS (R$)": resultado["ibs"],
                        "Imposto Bruto (R$)": resultado["imposto_bruto"],
                        "Créditos (R$)": resultado["creditos"],
                        "Imposto Devido (R$)": resultado["imposto_devido"],
                        "Carga Atual (R$)": valor_atual,
                        "Diferença (R$)": diferenca,
                        "Alíquota Efetiva (%)": resultado["aliquota_efetiva"] * 100
                    })

                # Ordenar dados por ano
                dados_tabela.sort(key=lambda x: x["Ano"])

                # Preencher tabela
                for i, dados in enumerate(dados_tabela, 4):
                    # Preencher dados sem formatação (para que os gráficos funcionem corretamente)
                    for col, cabecalho in enumerate(cabecalhos, 1):
                        valor = dados[cabecalho]
                        cell = ws_resultados.cell(row=i, column=col, value=valor)

                        # Formatação especial para colunas de valores
                        if col > 1 and col < 9:  # Colunas de valores monetários
                            cell.number_format = '#,##0.00'
                        elif col == 9:  # Coluna de percentual
                            cell.number_format = '0.00%'

                        # Destacar diferenças
                        if col == 8:  # Coluna de diferença
                            if valor > 0:
                                cell.font = Font(color="FF0000")  # Vermelho se aumentar
                            elif valor < 0:
                                cell.font = Font(color="008000")  # Verde se diminuir

                # Ajustar largura das colunas
                for col in range(1, len(cabecalhos) + 1):
                    ws_resultados.column_dimensions[chr(64 + col)].width = 18

                # Aba com Gráficos
                ws_graficos = wb.create_sheet(title="Gráficos")

                # Título da aba de gráficos
                ws_graficos['A1'] = "Análise Gráfica dos Resultados"
                ws_graficos['A1'].font = Font(bold=True, size=14)
                ws_graficos.merge_cells('A1:F1')

                # Criar gráfico de barras para comparação CBS/IBS
                chart1 = BarChart()
                chart1.title = "Comparativo: CBS vs IBS"
                chart1.style = 10
                chart1.type = "col"
                chart1.grouping = "clustered"
                chart1.y_axis.title = "Valor (R$)"
                chart1.x_axis.title = "Ano"

                data = Reference(ws_resultados, min_col=2, max_col=3, min_row=3, max_row=3 + len(dados_tabela))
                cats = Reference(ws_resultados, min_col=1, min_row=4, max_row=3 + len(dados_tabela))

                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(cats)
                chart1.height = 15
                chart1.width = 20

                # Adicionar este gráfico na planilha
                ws_graficos.add_chart(chart1, "A3")

                # Criar gráfico de comparação entre imposto devido e carga atual
                chart2 = BarChart()
                chart2.title = "Comparativo: Imposto Devido vs. Carga Atual"
                chart2.style = 11
                chart2.type = "col"
                chart2.grouping = "clustered"
                chart2.y_axis.title = "Valor (R$)"
                chart2.x_axis.title = "Ano"

                data = Reference(ws_resultados, min_col=6, max_col=7, min_row=3, max_row=3 + len(dados_tabela))

                chart2.add_data(data, titles_from_data=True)
                chart2.set_categories(cats)
                chart2.height = 15
                chart2.width = 20

                # Adicionar na planilha
                ws_graficos.add_chart(chart2, "A20")

                # Criar gráfico de linha para a alíquota efetiva
                chart3 = LineChart()
                chart3.title = "Evolução da Alíquota Efetiva"
                chart3.style = 12
                chart3.y_axis.title = "Alíquota (%)"
                chart3.x_axis.title = "Ano"

                data = Reference(ws_resultados, min_col=9, max_col=9, min_row=3, max_row=3 + len(dados_tabela))

                chart3.add_data(data, titles_from_data=True)
                chart3.set_categories(cats)
                chart3.height = 15
                chart3.width = 20

                # Adicionar rótulos de dados
                chart3.dataLabels = DataLabelList()
                chart3.dataLabels.showVal = True

                # Adicionar na planilha
                ws_graficos.add_chart(chart3, "H3")

                # Aba de Incentivos Fiscais
                ws_incentivos = wb.create_sheet(title="Incentivos Fiscais")

                # Título da aba de incentivos
                ws_incentivos['A1'] = "Incentivos Fiscais Configurados"
                ws_incentivos['A1'].font = Font(bold=True, size=14)
                ws_incentivos.merge_cells('A1:D1')

                # Incentivos de Saída
                ws_incentivos['A3'] = "Incentivos de Saída"
                ws_incentivos['A3'].font = Font(bold=True)
                ws_incentivos.merge_cells('A3:D3')

                # Cabeçalho dos incentivos
                cabecalhos_inc = ["Descrição", "Tipo", "Percentual", "% Operações"]
                for col, header in enumerate(cabecalhos_inc, 1):
                    cell = ws_incentivos.cell(row=4, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Preencher dados dos incentivos de saída
                row = 5
                for i, incentivo in enumerate(st.session_state.config.icms_config["incentivos_saida"]):
                    ws_incentivos.cell(row=row, column=1, value=incentivo["descricao"])
                    ws_incentivos.cell(row=row, column=2, value=incentivo["tipo"])
                    ws_incentivos.cell(row=row, column=3, value=incentivo["percentual"])
                    ws_incentivos.cell(row=row, column=4, value=incentivo["percentual_operacoes"])

                    # Formatar células
                    ws_incentivos.cell(row=row, column=3).number_format = '0.00%'
                    ws_incentivos.cell(row=row, column=4).number_format = '0.00%'

                    row += 1

                # Preencher incentivos de entrada
                row += 2
                ws_incentivos.cell(row=row, column=1, value="Incentivos de Entrada")
                ws_incentivos.cell(row=row, column=1).font = Font(bold=True)
                ws_incentivos.merge_cells(f'A{row}:D{row}')

                row += 1
                for col, header in enumerate(cabecalhos_inc, 1):
                    cell = ws_incentivos.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                row += 1
                for i, incentivo in enumerate(st.session_state.config.icms_config["incentivos_entrada"]):
                    ws_incentivos.cell(row=row, column=1, value=incentivo["descricao"])
                    ws_incentivos.cell(row=row, column=2, value=incentivo["tipo"])
                    ws_incentivos.cell(row=row, column=3, value=incentivo["percentual"])
                    ws_incentivos.cell(row=row, column=4, value=incentivo["percentual_operacoes"])

                    # Formatar células
                    ws_incentivos.cell(row=row, column=3).number_format = '0.00%'
                    ws_incentivos.cell(row=row, column=4).number_format = '0.00%'

                    row += 1

                # Preencher incentivos de apuração
                row += 2
                ws_incentivos.cell(row=row, column=1, value="Incentivos de Apuração")
                ws_incentivos.cell(row=row, column=1).font = Font(bold=True)
                ws_incentivos.merge_cells(f'A{row}:D{row}')

                row += 1
                for col, header in enumerate(cabecalhos_inc, 1):
                    cell = ws_incentivos.cell(row=row, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                row += 1
                for i, incentivo in enumerate(st.session_state.config.icms_config["incentivos_apuracao"]):
                    ws_incentivos.cell(row=row, column=1, value=incentivo["descricao"])
                    ws_incentivos.cell(row=row, column=2, value=incentivo["tipo"])
                    ws_incentivos.cell(row=row, column=3, value=incentivo["percentual"])
                    ws_incentivos.cell(row=row, column=4, value=incentivo["percentual_operacoes"])

                    # Formatar células
                    ws_incentivos.cell(row=row, column=3).number_format = '0.00%'
                    ws_incentivos.cell(row=row, column=4).number_format = '0.00%'

                    row += 1

                # Ajustar largura das colunas
                ws_incentivos.column_dimensions['A'].width = 30
                ws_incentivos.column_dimensions['B'].width = 25
                ws_incentivos.column_dimensions['C'].width = 15
                ws_incentivos.column_dimensions['D'].width = 15

                # Aba de Memória de Cálculo
                ws_memoria = wb.create_sheet(title="Memória de Cálculo")

                # Título da aba
                ws_memoria['A1'] = "Memória de Cálculo Detalhada"
                ws_memoria['A1'].font = Font(bold=True, size=14)
                ws_memoria.merge_cells('A1:E1')

                # Selecionar o primeiro ano disponível para a memória de cálculo
                if st.session_state.resultados:
                    primeiro_ano = sorted(st.session_state.resultados.keys())[0]

                    # Adicionar título do ano
                    ws_memoria['A3'] = f"Memória de Cálculo - Ano {primeiro_ano}"
                    ws_memoria['A3'].font = Font(bold=True)
                    ws_memoria.merge_cells('A3:E3')

                    # Obter a memória de cálculo
                    memoria = st.session_state.memoria_calculo

                    # Função auxiliar para adicionar seções
                    def adicionar_secao(titulo, secao_info, linha_inicio):
                        linha = linha_inicio

                        # Adicionar título da seção
                        ws_memoria.cell(row=linha, column=1, value=titulo)
                        ws_memoria.cell(row=linha, column=1).font = Font(bold=True)
                        ws_memoria.merge_cells(f'A{linha}:E{linha}')

                        linha += 1

                        # Adicionar conteúdo
                        if isinstance(secao_info, list):
                            # Caso de caminhos aninhados
                            if len(secao_info) >= 2 and secao_info[0] in memoria and secao_info[1] in memoria[
                                secao_info[0]]:
                                dados_secao = memoria[secao_info[0]][secao_info[1]]
                                if dados_secao:
                                    for item in dados_secao:
                                        ws_memoria.cell(row=linha, column=1, value=item)
                                        ws_memoria.merge_cells(f'A{linha}:E{linha}')
                                        linha += 1
                                else:
                                    ws_memoria.cell(row=linha, column=1, value="Não disponível")
                                    ws_memoria.merge_cells(f'A{linha}:E{linha}')
                                    linha += 1
                            else:
                                ws_memoria.cell(row=linha, column=1, value="Dados não disponíveis para esta seção")
                                ws_memoria.merge_cells(f'A{linha}:E{linha}')
                                linha += 1
                        else:
                            # Caso de chave direta
                            if secao_info in memoria and memoria[secao_info]:
                                for item in memoria[secao_info]:
                                    ws_memoria.cell(row=linha, column=1, value=item)
                                    ws_memoria.merge_cells(f'A{linha}:E{linha}')
                                    linha += 1
                            else:
                                ws_memoria.cell(row=linha, column=1, value="Não disponível")
                                ws_memoria.merge_cells(f'A{linha}:E{linha}')
                                linha += 1

                        linha += 1
                        return linha

                    # Adicionar seções da memória de cálculo
                    linha_atual = 5
                    linha_atual = adicionar_secao("VALIDAÇÃO DE DADOS", "validacao", linha_atual)
                    linha_atual = adicionar_secao("BASE TRIBUTÁVEL", "base_tributavel", linha_atual)
                    linha_atual = adicionar_secao("ALÍQUOTAS", "aliquotas", linha_atual)
                    linha_atual = adicionar_secao("CÁLCULO DA CBS", "cbs", linha_atual)
                    linha_atual = adicionar_secao("CÁLCULO DO IBS", "ibs", linha_atual)
                    linha_atual = adicionar_secao("CÁLCULO DOS CRÉDITOS", "creditos", linha_atual)
                    linha_atual = adicionar_secao("CÁLCULO DO IMPOSTO DEVIDO", "imposto_devido", linha_atual)

                    # Calcular os impostos atuais
                    ws_memoria.cell(row=linha_atual, column=1, value="CÁLCULO DOS IMPOSTOS ATUAIS")
                    ws_memoria.cell(row=linha_atual, column=1).font = Font(bold=True, size=12)
                    ws_memoria.merge_cells(f'A{linha_atual}:E{linha_atual}')
                    linha_atual += 2

                    if "impostos_atuais" in memoria:
                        linha_atual = adicionar_secao("PIS", ["impostos_atuais", "PIS"], linha_atual)
                        linha_atual = adicionar_secao("COFINS", ["impostos_atuais", "COFINS"], linha_atual)
                        linha_atual = adicionar_secao("ICMS", ["impostos_atuais", "ICMS"], linha_atual)
                        linha_atual = adicionar_secao("ISS", ["impostos_atuais", "ISS"], linha_atual)
                        linha_atual = adicionar_secao("IPI", ["impostos_atuais", "IPI"], linha_atual)
                        linha_atual = adicionar_secao("TOTAL IMPOSTOS ATUAIS", ["impostos_atuais", "total"],
                                                      linha_atual)

                    # Adicionar créditos cruzados e total devido
                    if "creditos_cruzados" in memoria and memoria["creditos_cruzados"]:
                        linha_atual = adicionar_secao("CRÉDITOS CRUZADOS", "creditos_cruzados", linha_atual)

                    linha_atual = adicionar_secao("TOTAL DEVIDO", "total_devido", linha_atual)

                # Ajustar largura das colunas
                ws_memoria.column_dimensions['A'].width = 100

                # Aba com Alíquotas Setoriais
                ws_setores = wb.create_sheet(title="Alíquotas Setoriais")

                # Título
                ws_setores['A1'] = "Alíquotas por Setor - LC 214/2025"
                ws_setores['A1'].font = Font(bold=True, size=14)
                ws_setores.merge_cells('A1:C1')

                # Cabeçalho
                cabecalhos = ["Setor", "IBS (%)", "Redução CBS (%)"]
                for col, header in enumerate(cabecalhos, 1):
                    cell = ws_setores.cell(row=3, column=col, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')

                # Dados das alíquotas setoriais
                row = 4
                for setor, valores in st.session_state.config.setores_especiais.items():
                    ws_setores.cell(row=row, column=1, value=setor)
                    ws_setores.cell(row=row, column=2, value=valores["IBS"])
                    ws_setores.cell(row=row, column=3, value=valores["reducao_CBS"])

                    # Formatação
                    ws_setores.cell(row=row, column=2).number_format = '0.00%'
                    ws_setores.cell(row=row, column=3).number_format = '0.00%'

                    row += 1

                # Ajustar largura das colunas
                ws_setores.column_dimensions['A'].width = 20
                ws_setores.column_dimensions['B'].width = 15
                ws_setores.column_dimensions['C'].width = 20

                # Rodapé em todas as abas
                for ws in wb.worksheets:
                    row = ws.max_row + 2
                    ws.cell(row=row, column=1, value="© 2025 Expertzy Inteligência Tributária")
                    ws.cell(row=row, column=1).font = Font(italic=True, size=9)

                # Salvar o arquivo
                wb.save(output)

                # Retornar o arquivo como base64 para download
                output.seek(0)
                b64 = base64.b64encode(output.read()).decode()
                return b64

            except Exception as e:
                import traceback
                st.error(f"Erro ao gerar Excel: {e}")
                st.error(traceback.format_exc())
                return None

        # Substituir no app.py a parte onde está o botão de exportação para Excel:
        if st.button("Exportar para Excel", key="bt_excel"):
            with st.spinner("Gerando arquivo Excel..."):
                b64_excel = exportar_excel()
                if b64_excel:
                    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}" download="simulacao_iva_dual.xlsx">Download do arquivo Excel</a>'
                    st.markdown(href, unsafe_allow_html=True)
        return output.getvalue()
    elif formato == "pdf":
        output = BytesIO()

        def exportar_pdf():
            """Exporta os resultados para um arquivo PDF detalhado."""
            if not st.session_state.resultados:
                st.warning("Execute uma simulação antes de exportar os resultados.")
                return None

            try:
                # Criar um buffer em memória para o PDF
                buffer = io.BytesIO()

                # Configuração do documento
                doc = SimpleDocTemplate(
                    buffer,
                    pagesize=letter,
                    rightMargin=72,
                    leftMargin=72,
                    topMargin=72,
                    bottomMargin=72
                )

                # Lista de elementos para o PDF
                elementos = []

                # Estilos
                estilos = getSampleStyleSheet()
                titulo_estilo = estilos['Heading1']
                subtitulo_estilo = estilos['Heading2']
                subsecao_estilo = estilos['Heading3']
                normal_estilo = estilos['Normal']

                # Criar estilo para código/memória de cálculo
                codigo_estilo = ParagraphStyle(
                    'CodigoEstilo',
                    parent=estilos['Normal'],
                    fontName='Courier',
                    fontSize=8,
                    leading=10,
                    leftIndent=36,
                )

                # Adicionar título
                elementos.append(Paragraph("Relatório de Simulação - IVA Dual (CBS/IBS)", titulo_estilo))
                elementos.append(Spacer(1, 0.25 * inch))

                # Data do relatório
                import datetime
                data_hora = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
                elementos.append(Paragraph(f"Data do relatório: {data_hora}", normal_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Parâmetros da simulação
                elementos.append(Paragraph("Parâmetros da Simulação", subtitulo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Obter dados da sessão
                faturamento = st.session_state.get('faturamento', 0)
                custos = st.session_state.get('custos_tributaveis', 0)
                custos_simples = st.session_state.get('custos_simples', 0)
                creditos_anteriores = st.session_state.get('creditos_anteriores', 0)
                setor = st.session_state.get('setor', 'padrao')
                regime = st.session_state.get('regime', 'real')
                carga_atual = st.session_state.get('carga_atual', 25)
                aliquota_entrada = st.session_state.get('aliquota_entrada', 19)
                aliquota_saida = st.session_state.get('aliquota_saida', 19)

                # Tabela com parâmetros
                dados_parametros = [
                    ["Parâmetro", "Valor"],
                    ["Faturamento Anual", f"R$ {formatar_br(faturamento)}"],
                    ["Custos Tributáveis", f"R$ {formatar_br(custos)}"],
                    ["Fornecedores do Simples", f"R$ {formatar_br(custos_simples)}"],
                    ["Créditos Anteriores", f"R$ {formatar_br(creditos_anteriores)}"],
                    ["Setor de Atividade", setor],
                    ["Regime Tributário", regime],
                    ["Carga Tributária Atual", f"{formatar_br(carga_atual)}%"],
                    ["Alíquota ICMS Entrada", f"{formatar_br(aliquota_entrada)}%"],
                    ["Alíquota ICMS Saída", f"{formatar_br(aliquota_saida)}%"]
                ]

                tabela_parametros = Table(dados_parametros, colWidths=[2.5 * inch, 2.5 * inch])
                tabela_parametros.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (1, 0), 12),
                    ('BACKGROUND', (0, 1), (1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))

                elementos.append(tabela_parametros)
                elementos.append(Spacer(1, 0.2 * inch))

                # Incentivos Fiscais
                elementos.append(Paragraph("Incentivos Fiscais Configurados", subtitulo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Incentivos de Saída
                incentivos_saida = st.session_state.config.icms_config["incentivos_saida"]
                if incentivos_saida:
                    elementos.append(Paragraph("Incentivos de Saída", subsecao_estilo))

                    dados_saida = [["Descrição", "Tipo", "Percentual", "% Operações"]]
                    for inc in incentivos_saida:
                        dados_saida.append([
                            inc["descricao"],
                            inc["tipo"],
                            f"{formatar_br(inc['percentual'] * 100)}%",
                            f"{formatar_br(inc['percentual_operacoes'] * 100)}%"
                        ])

                    tabela_saida = Table(dados_saida, colWidths=[2 * inch, 2 * inch, 1 * inch, 1 * inch])
                    tabela_saida.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]))

                    elementos.append(tabela_saida)
                    elementos.append(Spacer(1, 0.2 * inch))
                else:
                    elementos.append(Paragraph("Nenhum incentivo de saída configurado.", normal_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                # Incentivos de Entrada
                incentivos_entrada = st.session_state.config.icms_config["incentivos_entrada"]
                if incentivos_entrada:
                    elementos.append(Paragraph("Incentivos de Entrada", subsecao_estilo))

                    dados_entrada = [["Descrição", "Tipo", "Percentual", "% Operações"]]
                    for inc in incentivos_entrada:
                        dados_entrada.append([
                            inc["descricao"],
                            inc["tipo"],
                            f"{formatar_br(inc['percentual'] * 100)}%",
                            f"{formatar_br(inc['percentual_operacoes'] * 100)}%"
                        ])

                    tabela_entrada = Table(dados_entrada, colWidths=[2 * inch, 2 * inch, 1 * inch, 1 * inch])
                    tabela_entrada.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]))

                    elementos.append(tabela_entrada)
                    elementos.append(Spacer(1, 0.2 * inch))
                else:
                    elementos.append(Paragraph("Nenhum incentivo de entrada configurado.", normal_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                # Incentivos de Apuração
                incentivos_apuracao = st.session_state.config.icms_config["incentivos_apuracao"]
                if incentivos_apuracao:
                    elementos.append(Paragraph("Incentivos de Apuração", subsecao_estilo))

                    dados_apuracao = [["Descrição", "Tipo", "Percentual", "% do Saldo"]]
                    for inc in incentivos_apuracao:
                        dados_apuracao.append([
                            inc["descricao"],
                            inc["tipo"],
                            f"{formatar_br(inc['percentual'] * 100)}%",
                            f"{formatar_br(inc['percentual_operacoes'] * 100)}%"
                        ])

                    tabela_apuracao = Table(dados_apuracao, colWidths=[2 * inch, 2 * inch, 1 * inch, 1 * inch])
                    tabela_apuracao.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black),
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ]))

                    elementos.append(tabela_apuracao)
                    elementos.append(Spacer(1, 0.2 * inch))
                else:
                    elementos.append(Paragraph("Nenhum incentivo de apuração configurado.", normal_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                # Colocar este código após adicionar os incentivos de apuração e antes de executar a simulação
                if st.session_state.config.icms_config["incentivos_apuracao"]:
                    st.success(
                        f"{len(st.session_state.config.icms_config['incentivos_apuracao'])} incentivo(s) de apuração configurado(s). Estes serão aplicados sobre o saldo devedor após a compensação de débitos e créditos.")

                # Adicionar após a simulação, dentro do bloco de exibição de resultados
                if "resultados" in st.session_state and st.session_state.resultados and \
                        st.session_state.config.icms_config["incentivos_apuracao"]:
                    # Obter o primeiro ano
                    primeiro_ano = min(st.session_state.resultados.keys())
                    resultado = st.session_state.resultados[primeiro_ano]

                    # Verificar se há economia de ICMS
                    if "economia_icms" in resultado["impostos_atuais"] and resultado["impostos_atuais"][
                        "economia_icms"] > 0:
                        st.info(f"Os incentivos de apuração configurados contribuíram para a economia fiscal. " +
                                f"O ICMS devido foi reduzido pela aplicação dos incentivos de apuração.")

                        # Adicionar uma métrica visual
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric(
                                label="ICMS sem Incentivos",
                                value=f"R$ {formatar_br(resultado['impostos_atuais']['ICMS'] + resultado['impostos_atuais']['economia_icms'])}"
                            )
                        with col2:
                            st.metric(
                                label="ICMS com Incentivos",
                                value=f"R$ {formatar_br(resultado['impostos_atuais']['ICMS'])}",
                                delta=f"-{formatar_br(resultado['impostos_atuais']['economia_icms'])}"
                            )
                        with col3:
                            # Calcular percentual de economia
                            perc_economia = (resultado['impostos_atuais']['economia_icms'] /
                                             (resultado['impostos_atuais']['ICMS'] + resultado['impostos_atuais'][
                                                 'economia_icms'])) * 100
                            st.metric(
                                label="Economia Percentual",
                                value=f"{formatar_br(perc_economia)}%"
                            )

                # Resultados da simulação
                elementos.append(Paragraph("Resultados da Simulação", subtitulo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Preparar dados da tabela de resultados
                dados_tabela = [
                    ["Ano", "CBS (R$)", "IBS (R$)", "Imposto Bruto (R$)", "Créditos (R$)",
                     "Imposto Devido (R$)", "Carga Atual (R$)", "Diferença (R$)", "Alíquota (%)"]
                ]

                # Ordenar resultados por ano
                anos_ordenados = sorted(st.session_state.resultados.keys())

                for ano in anos_ordenados:
                    resultado = st.session_state.resultados[ano]
                    valor_atual = st.session_state.aliquotas_equivalentes[ano]["valor_atual"]
                    diferenca = resultado["imposto_devido"] - valor_atual

                    dados_tabela.append([
                        str(ano),
                        f"R$ {formatar_br(resultado['cbs'])}",
                        f"R$ {formatar_br(resultado['ibs'])}",
                        f"R$ {formatar_br(resultado['imposto_bruto'])}",
                        f"R$ {formatar_br(resultado['creditos'])}",
                        f"R$ {formatar_br(resultado['imposto_devido'])}",
                        f"R$ {formatar_br(valor_atual)}",
                        f"R$ {formatar_br(diferenca)}",
                        f"{formatar_br(resultado['aliquota_efetiva'] * 100)}%"
                    ])

                # Criar a tabela de resultados
                tabela_resultados = Table(dados_tabela, colWidths=[0.5 * inch, 0.85 * inch, 0.85 * inch,
                                                                   0.9 * inch, 0.85 * inch, 0.9 * inch,
                                                                   0.85 * inch, 0.85 * inch, 0.7 * inch])

                tabela_resultados.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ]))

                elementos.append(tabela_resultados)
                elementos.append(Spacer(1, 0.2 * inch))

                # Análise dos resultados
                elementos.append(Paragraph("Análise dos Resultados", subtitulo_estilo))
                elementos.append(Spacer(1, 0.1 * inch))

                # Adicionar texto de análise
                if len(anos_ordenados) >= 2:
                    ano_inicial = min(anos_ordenados)
                    ano_final = max(anos_ordenados)

                    resultado_inicial = st.session_state.resultados[ano_inicial]
                    resultado_final = st.session_state.resultados[ano_final]

                    # Variação do imposto devido
                    variacao_imposto = resultado_final["imposto_devido"] - resultado_inicial["imposto_devido"]
                    percentual_var = (variacao_imposto / resultado_inicial["imposto_devido"]) * 100 if \
                    resultado_inicial["imposto_devido"] > 0 else 0

                    # Variação da alíquota efetiva
                    variacao_aliquota = resultado_final["aliquota_efetiva"] - resultado_inicial["aliquota_efetiva"]
                    percentual_var_aliq = variacao_aliquota * 100

                    texto_analise = f"""
                    Durante o período de transição do IVA Dual (de {ano_inicial} a {ano_final}), observa-se uma 
                    variação significativa na carga tributária da empresa. O imposto devido no IVA Dual passa 
                    de R$ {formatar_br(resultado_inicial["imposto_devido"])} para R$ {formatar_br(resultado_final["imposto_devido"])}, 
                    representando uma variação de {formatar_br(percentual_var)}%.

                    A alíquota efetiva evolui de {formatar_br(resultado_inicial["aliquota_efetiva"] * 100)}% para 
                    {formatar_br(resultado_final["aliquota_efetiva"] * 100)}%, uma variação de {formatar_br(percentual_var_aliq)} pontos percentuais.

                    Esta evolução reflete a implementação progressiva do novo sistema tributário, conforme 
                    estabelecido pela Lei Complementar 214/2025. Vale destacar que o impacto da reforma varia de 
                    acordo com o setor de atividade, o regime tributário e a estrutura de custos da empresa.
                    """

                    elementos.append(Paragraph(texto_analise.strip(), normal_estilo))
                    elementos.append(Spacer(1, 0.2 * inch))

                # Adicionar comentário sobre ICMS e incentivos fiscais
                if st.session_state.config.icms_config["incentivos_saida"] or st.session_state.config.icms_config[
                    "incentivos_entrada"] or st.session_state.config.icms_config["incentivos_apuracao"]:
                    resultado = st.session_state.resultados[anos_ordenados[0]]
                    icms_devido = resultado["impostos_atuais"].get("ICMS", 0)
                    economia_icms = resultado["impostos_atuais"].get("economia_icms", 0)
                    total_incentivos = len(st.session_state.config.icms_config["incentivos_saida"]) + len(
                        st.session_state.config.icms_config["incentivos_entrada"]) + len(
                        st.session_state.config.icms_config["incentivos_apuracao"])

                    texto_incentivos = f"""
                    A simulação considera {total_incentivos} incentivos fiscais configurados, resultando em uma 
                    economia fiscal de R$ {formatar_br(economia_icms)} no ICMS. Sem estes incentivos, o ICMS devido 
                    seria de R$ {formatar_br(icms_devido + economia_icms)}, em vez dos atuais R$ {formatar_br(icms_devido)}.

                    Durante a transição para o IVA Dual, é importante considerar como estes incentivos serão tratados, 
                    pois a reforma tributária pode afetar significativamente os benefícios fiscais existentes.
                    """

                    elementos.append(Paragraph("Impacto dos Incentivos Fiscais", subsecao_estilo))
                    elementos.append(Paragraph(texto_incentivos.strip(), normal_estilo))
                    elementos.append(Spacer(1, 0.2 * inch))

                # Memória de Cálculo
                elementos.append(PageBreak())
                elementos.append(Paragraph("Memória de Cálculo", titulo_estilo))
                elementos.append(Spacer(1, 0.25 * inch))

                # Selecionar primeiro ano para demonstração
                if st.session_state.resultados:
                    primeiro_ano = min(anos_ordenados)
                    memoria = st.session_state.memoria_calculo

                    # Validação de dados
                    elementos.append(Paragraph("Validação de Dados", subtitulo_estilo))
                    if "validacao" in memoria and memoria["validacao"]:
                        for linha in memoria["validacao"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    else:
                        elementos.append(Paragraph("Dados validados com sucesso.", codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # Base tributável
                    elementos.append(Paragraph("Base Tributável", subtitulo_estilo))
                    if "base_tributavel" in memoria and memoria["base_tributavel"]:
                        for linha in memoria["base_tributavel"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # Alíquotas
                    elementos.append(Paragraph("Alíquotas", subtitulo_estilo))
                    if "aliquotas" in memoria and memoria["aliquotas"]:
                        for linha in memoria["aliquotas"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # Cálculo da CBS
                    elementos.append(Paragraph("Cálculo da CBS", subtitulo_estilo))
                    if "cbs" in memoria and memoria["cbs"]:
                        for linha in memoria["cbs"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # Cálculo do IBS
                    elementos.append(Paragraph("Cálculo do IBS", subtitulo_estilo))
                    if "ibs" in memoria and memoria["ibs"]:
                        for linha in memoria["ibs"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # Nova página para o resto da memória
                    elementos.append(PageBreak())

                    # Cálculo dos Créditos
                    elementos.append(Paragraph("Cálculo dos Créditos", subtitulo_estilo))
                    if "creditos" in memoria and memoria["creditos"]:
                        for linha in memoria["creditos"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # Cálculo do Imposto Devido
                    elementos.append(Paragraph("Cálculo do Imposto Devido", subtitulo_estilo))
                    if "imposto_devido" in memoria and memoria["imposto_devido"]:
                        for linha in memoria["imposto_devido"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                    # Cálculo dos Impostos Atuais
                    elementos.append(PageBreak())
                    elementos.append(Paragraph("Cálculo dos Impostos Atuais", titulo_estilo))
                    elementos.append(Spacer(1, 0.25 * inch))

                    if "impostos_atuais" in memoria:
                        # PIS
                        elementos.append(Paragraph("PIS", subtitulo_estilo))
                        if "PIS" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["PIS"]:
                            for linha in memoria["impostos_atuais"]["PIS"]:
                                elementos.append(Paragraph(linha, codigo_estilo))
                        elementos.append(Spacer(1, 0.1 * inch))

                        # COFINS
                        elementos.append(Paragraph("COFINS", subtitulo_estilo))
                        if "COFINS" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["COFINS"]:
                            for linha in memoria["impostos_atuais"]["COFINS"]:
                                elementos.append(Paragraph(linha, codigo_estilo))
                        elementos.append(Spacer(1, 0.1 * inch))

                        # ICMS (mais detalhado)
                        elementos.append(Paragraph("ICMS", subtitulo_estilo))
                        if "ICMS" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["ICMS"]:
                            for linha in memoria["impostos_atuais"]["ICMS"]:
                                elementos.append(Paragraph(linha, codigo_estilo))
                        elementos.append(Spacer(1, 0.1 * inch))

                        elementos.append(PageBreak())

                        # ISS
                        elementos.append(Paragraph("ISS", subtitulo_estilo))
                        if "ISS" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["ISS"]:
                            for linha in memoria["impostos_atuais"]["ISS"]:
                                elementos.append(Paragraph(linha, codigo_estilo))
                        elementos.append(Spacer(1, 0.1 * inch))

                        # IPI
                        elementos.append(Paragraph("IPI", subtitulo_estilo))
                        if "IPI" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["IPI"]:
                            for linha in memoria["impostos_atuais"]["IPI"]:
                                elementos.append(Paragraph(linha, codigo_estilo))
                        elementos.append(Spacer(1, 0.1 * inch))

                        # Total Impostos Atuais
                        elementos.append(Paragraph("Total Impostos Atuais", subtitulo_estilo))
                        if "total" in memoria["impostos_atuais"] and memoria["impostos_atuais"]["total"]:
                            for linha in memoria["impostos_atuais"]["total"]:
                                elementos.append(Paragraph(linha, codigo_estilo))
                        elementos.append(Spacer(1, 0.1 * inch))

                    # Créditos Cruzados
                    if "creditos_cruzados" in memoria and memoria["creditos_cruzados"]:
                        elementos.append(Paragraph("Créditos Cruzados", subtitulo_estilo))
                        for linha in memoria["creditos_cruzados"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                        elementos.append(Spacer(1, 0.1 * inch))

                    # Total Devido
                    elementos.append(Paragraph("Total Devido", subtitulo_estilo))
                    if "total_devido" in memoria and memoria["total_devido"]:
                        for linha in memoria["total_devido"]:
                            elementos.append(Paragraph(linha, codigo_estilo))
                    elementos.append(Spacer(1, 0.1 * inch))

                # Rodapé
                elementos.append(Spacer(1, 0.5 * inch))
                elementos.append(Paragraph("© 2025 Expertzy Inteligência Tributária",
                                           ParagraphStyle('rodape',
                                                          parent=normal_estilo,
                                                          alignment=1,  # centralizado
                                                          fontSize=8,
                                                          textColor=colors.darkgrey)))

                # Construir o PDF
                doc.build(elementos)

                # Retornar o conteúdo do buffer
                buffer.seek(0)
                b64 = base64.b64encode(buffer.read()).decode()
                return b64

            except Exception as e:
                import traceback
                st.error(f"Erro ao gerar PDF: {e}")
                st.error(traceback.format_exc())
                return None

        # Substituir no app.py a parte onde está o botão de exportação para PDF:
        if st.button("Exportar para PDF", key="bt_pdf"):
            with st.spinner("Gerando arquivo PDF..."):
                b64_pdf = exportar_pdf()
                if b64_pdf:
                    href = f'<a href="data:application/pdf;base64,{b64_pdf}" download="simulacao_iva_dual.pdf">Download do arquivo PDF</a>'
                    st.markdown(href, unsafe_allow_html=True)
        return output.getvalue()

    return None


# Função para executar a simulação
# Adicionar ao arquivo app.py - Correção para o executar_simulacao para armazenar mais dados

def executar_simulacao(dados_empresa, ano_inicial, ano_final):
    """Executa a simulação com os dados fornecidos e armazena mais informações."""
    # Armazenar dados no estado da sessão para exportação posterior
    st.session_state.faturamento = dados_empresa.get("faturamento", 0)
    st.session_state.custos_tributaveis = dados_empresa.get("custos_tributaveis", 0)
    st.session_state.custos_simples = dados_empresa.get("custos_simples", 0)
    st.session_state.creditos_anteriores = dados_empresa.get("creditos_anteriores", 0)
    st.session_state.setor = dados_empresa.get("setor", "padrao")
    st.session_state.regime = dados_empresa.get("regime", "real")
    st.session_state.carga_atual = dados_empresa.get("carga_atual", 25)
    st.session_state.aliquota_entrada = dados_empresa.get("aliquota_entrada", 19)
    st.session_state.aliquota_saida = dados_empresa.get("aliquota_saida", 19)

    # Atualizar configurações do ICMS
    st.session_state.config.icms_config["aliquota_entrada"] = dados_empresa.get("aliquota_entrada", 19) / 100
    st.session_state.config.icms_config["aliquota_saida"] = dados_empresa.get("aliquota_saida", 19) / 100

    # Preparar dados para a simulação
    dados_simulacao = {
        "faturamento": dados_empresa.get("faturamento", 0),
        "custos_tributaveis": dados_empresa.get("custos_tributaveis", 0),
        "custos_simples": dados_empresa.get("custos_simples", 0),
        "creditos_anteriores": dados_empresa.get("creditos_anteriores", 0),
        "setor": dados_empresa.get("setor", "padrao"),
        "regime": dados_empresa.get("regime", "real"),
        "imposto_devido": 0  # Será calculado iterativamente
    }

    # Definir anos para simulação
    anos = list(range(ano_inicial, ano_final + 1))

    try:
        # Executar simulação
        resultados = st.session_state.calculadora_iva.calcular_comparativo(dados_simulacao, anos)
        st.session_state.resultados = resultados

        # Calcular alíquotas equivalentes
        carga_atual = dados_empresa.get("carga_atual", 25)
        aliquotas_equivalentes = {}

        for ano in anos:
            aliquotas_equivalentes[ano] = st.session_state.calculadora_iva.calcular_aliquotas_equivalentes(
                dados_simulacao, carga_atual, ano
            )

        st.session_state.aliquotas_equivalentes = aliquotas_equivalentes

        # Obter memória de cálculo
        st.session_state.memoria_calculo = st.session_state.calculadora_iva.memoria_calculo

        # Verificar dados específicos para incentivos de apuração
        for ano, resultado in resultados.items():
            # Verificar se há incentivos de apuração e se eles estão sendo aplicados
            if "economia_icms" in resultado["impostos_atuais"] and len(
                    st.session_state.config.icms_config["incentivos_apuracao"]) > 0:
                # Armazenar informações sobre economia específica de incentivos de apuração
                # (Isto seria mais preciso se implementado completamente na classe de cálculo)
                st.session_state[f"economia_apuracao_{ano}"] = resultado["impostos_atuais"][
                                                                   "economia_icms"] * 0.2  # Estimativa

        return True
    except Exception as e:
        st.error(f"Erro na simulação: {str(e)}")
        return False


# Inicializar sessão
inicializar_sessao()

# Sidebar para configurações e ações
st.sidebar.title("Simulador da Reforma Tributária")
st.sidebar.image(
    "https://www.gov.br/receitafederal/pt-br/assuntos/noticias/2022/dezembro/reforma-tributaria/art-reforma-jpg/@@images/image.jpeg",
    use_container_width=True)
opcao_sidebar = st.sidebar.radio("Navegação", ["Simulação", "Configurações", "Memória de Cálculo", "Sobre"])

# Conteúdo principal
if opcao_sidebar == "Simulação":
    st.title("Simulação do IVA Dual (CBS/IBS)")

    # Dividir a tela em colunas
    col1, col2 = st.columns([1, 2])

    with col1:
        st.subheader("Dados da Empresa")

        # Formulário para entrada de dados
        with st.form(key="form_dados_empresa"):
            faturamento = st.number_input("Faturamento Anual (R$)", min_value=0.0, value=1000000.0, step=10000.0,
                                          format="%.2f")
            custos = st.number_input("Custos Tributáveis (R$)", min_value=0.0, value=400000.0, step=10000.0,
                                     format="%.2f")
            custos_simples = st.number_input("Custos de Fornecedores do Simples (R$)", min_value=0.0, value=0.0,
                                             step=10000.0, format="%.2f")
            creditos_anteriores = st.number_input("Créditos Anteriores (R$)", min_value=0.0, value=0.0, step=10000.0,
                                                  format="%.2f")

            setor = st.selectbox("Setor de Atividade", list(st.session_state.config.setores_especiais.keys()))
            regime = st.selectbox("Regime Tributário", ["real", "presumido", "simples"])

            # Parâmetros de ICMS
            st.subheader("Parâmetros de ICMS")

            aliquota_entrada = st.number_input("Alíquota Média de Entrada (%)", min_value=0.0, max_value=100.0,
                                               value=19.0, step=0.1, format="%.2f")
            aliquota_saida = st.number_input("Alíquota Média de Saída (%)", min_value=0.0, max_value=100.0, value=19.0,
                                             step=0.1, format="%.2f")

            # Configurar período de simulação
            st.subheader("Período da Simulação")

            ano_inicial = st.selectbox("Ano Inicial", list(st.session_state.config.fase_transicao.keys()))
            ano_final = st.selectbox("Ano Final", list(st.session_state.config.fase_transicao.keys()),
                                     index=len(st.session_state.config.fase_transicao.keys()) - 1)

            # Carga tributária atual
            carga_atual = st.number_input("Carga Tributária Atual Estimada (%)", min_value=0.0, max_value=100.0,
                                          value=25.0, step=0.5, format="%.2f")

            # Botão para simular
            simular = st.form_submit_button("Simular")

        # Processar simulação se o botão for clicado
        if simular:
            dados_empresa = {
                "faturamento": faturamento,
                "custos_tributaveis": custos,
                "custos_simples": custos_simples,
                "creditos_anteriores": creditos_anteriores,
                "setor": setor,
                "regime": regime,
                "aliquota_entrada": aliquota_entrada,
                "aliquota_saida": aliquota_saida,
                "carga_atual": carga_atual
            }

            with st.spinner("Executando simulação..."):
                sucesso = executar_simulacao(dados_empresa, ano_inicial, ano_final)

                if sucesso:
                    st.success("Simulação concluída com sucesso!")

        # Incentivos fiscais (expander)
        with st.expander("Incentivos Fiscais de ICMS"):
            # Tabs para separar os tipos de incentivos
            tab_saida, tab_entrada, tab_apuracao = st.tabs(
                ["Incentivos de Saída", "Incentivos de Entrada", "Incentivos de Apuração"])

            # Incentivos de Saída
            with tab_saida:
                st.subheader("Incentivos de Saída")

                # Listar incentivos existentes
                if st.session_state.config.icms_config["incentivos_saida"]:
                    st.write("Incentivos configurados:")
                    for i, incentivo in enumerate(st.session_state.config.icms_config["incentivos_saida"]):
                        col_inc1, col_inc2, col_inc3 = st.columns([3, 1, 1])
                        with col_inc1:
                            st.write(f"{incentivo['descricao']} ({incentivo['tipo']})")
                        with col_inc2:
                            st.write(f"{formatar_br(incentivo['percentual'] * 100)}%")
                        with col_inc3:
                            if st.button(f"Remover#s{i}", key=f"remover_saida_{i}"):
                                remover_incentivo("saida", i)
                                st.rerun()

                # Formulário para adicionar novo incentivo
                with st.form(key="form_incentivo_saida"):
                    st.subheader("Adicionar Incentivo de Saída")

                    desc_saida = st.text_input("Descrição", value="", key="desc_saida")
                    tipo_saida = st.selectbox(
                        "Tipo",
                        ["Nenhum", "Redução de Alíquota", "Crédito Presumido/Outorgado", "Redução de Base de Cálculo",
                         "Diferimento"],
                        key="tipo_saida"
                    )
                    perc_saida = st.number_input("Percentual do Incentivo (%)", min_value=0.0, max_value=100.0,
                                                 value=0.0, step=1.0, key="perc_saida")
                    oper_saida = st.number_input("Percentual de Operações (%)", min_value=0.0, max_value=100.0,
                                                 value=100.0, step=1.0, key="oper_saida")
                    adicionar_saida = st.form_submit_button("Adicionar")

                if adicionar_saida:
                    dados_incentivo = {
                        "descricao": desc_saida or f"Incentivo Saída {len(st.session_state.config.icms_config['incentivos_saida']) + 1}",
                        "tipo": tipo_saida,
                        "percentual": perc_saida,
                        "perc_operacoes": oper_saida
                    }
                    if adicionar_incentivo("saida", dados_incentivo):
                        st.success("Incentivo de saída adicionado com sucesso!")
                        st.rerun()

            # Incentivos de Entrada
            with tab_entrada:
                st.subheader("Incentivos de Entrada")

                # Listar incentivos existentes
                if st.session_state.config.icms_config["incentivos_entrada"]:
                    st.write("Incentivos configurados:")
                    for i, incentivo in enumerate(st.session_state.config.icms_config["incentivos_entrada"]):
                        col_inc1, col_inc2, col_inc3 = st.columns([3, 1, 1])
                        with col_inc1:
                            st.write(f"{incentivo['descricao']} ({incentivo['tipo']})")
                        with col_inc2:
                            st.write(f"{formatar_br(incentivo['percentual'] * 100)}%")
                        with col_inc3:
                            if st.button(f"Remover#e{i}", key=f"remover_entrada_{i}"):
                                remover_incentivo("entrada", i)
                                st.rerun()

                # Formulário para adicionar novo incentivo
                with st.form(key="form_incentivo_entrada"):
                    st.subheader("Adicionar Incentivo de Entrada")

                    desc_entrada = st.text_input("Descrição", value="", key="desc_entrada")
                    tipo_entrada = st.selectbox(
                        "Tipo",
                        ["Nenhum", "Redução de Alíquota", "Crédito Presumido/Outorgado", "Redução de Base de Cálculo", "Estorno de Crédito"],
                        key="tipo_entrada"
                    )
                    perc_entrada = st.number_input("Percentual do Incentivo (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0, key="perc_entrada")
                    oper_entrada = st.number_input("Percentual de Operações (%)", min_value=0.0, max_value=100.0, value=100.0, step=1.0, key="oper_entrada")

                    adicionar_entrada = st.form_submit_button("Adicionar")

                if adicionar_entrada:
                    dados_incentivo = {
                        "descricao": desc_entrada or f"Incentivo Entrada {len(st.session_state.config.icms_config['incentivos_entrada']) + 1}",
                        "tipo": tipo_entrada,
                        "percentual": perc_entrada,
                        "perc_operacoes": oper_entrada
                    }
                    if adicionar_incentivo("entrada", dados_incentivo):
                        st.success("Incentivo de entrada adicionado com sucesso!")
                        st.rerun()

            # Incentivos de Apuração
            with tab_apuracao:
                st.subheader("Incentivos de Apuração")

                # Listar incentivos existentes
                if st.session_state.config.icms_config["incentivos_apuracao"]:
                    st.write("Incentivos configurados:")
                    for i, incentivo in enumerate(st.session_state.config.icms_config["incentivos_apuracao"]):
                        col_inc1, col_inc2, col_inc3 = st.columns([3, 1, 1])
                        with col_inc1:
                            st.write(f"{incentivo['descricao']} ({incentivo['tipo']})")
                        with col_inc2:
                            st.write(f"{formatar_br(incentivo['percentual'] * 100)}%")
                        with col_inc3:
                            if st.button(f"Remover#a{i}", key=f"remover_apuracao_{i}"):
                                remover_incentivo("apuracao", i)
                                st.rerun()

                # Formulário para adicionar novo incentivo
                with st.form(key="form_incentivo_apuracao"):
                    st.subheader("Adicionar Incentivo de Apuração")

                    desc_apuracao = st.text_input("Descrição", value="", key="desc_apuracao")
                    tipo_apuracao = st.selectbox(
                        "Tipo",
                        ["Nenhum", "Crédito Presumido/Outorgado", "Redução do Saldo Devedor"],
                        key="tipo_apuracao"
                    )
                    perc_apuracao = st.number_input("Percentual do Incentivo (%)", min_value=0.0, max_value=100.0, value=0.0, step=1.0, key="perc_apuracao")
                    oper_apuracao = st.number_input("Percentual do Saldo (%)", min_value=0.0, max_value=100.0, value=100.0, step=1.0, key="oper_apuracao")

                    adicionar_apuracao = st.form_submit_button("Adicionar")

                if adicionar_apuracao:
                    dados_incentivo = {
                        "descricao": desc_apuracao or f"Incentivo Apuração {len(st.session_state.config.icms_config['incentivos_apuracao']) + 1}",
                        "tipo": tipo_apuracao,
                        "percentual": perc_apuracao,
                        "perc_operacoes": oper_apuracao
                    }
                    if adicionar_incentivo("apuracao", dados_incentivo):
                        st.success("Incentivo de apuração adicionado com sucesso!")
                        st.rerun()

    # Coluna para exibição dos resultados
    with col2:
        if st.session_state.resultados:
            # Exibir tabela de resultados
            st.subheader("Resultados da Simulação")

            # Preparar dados para a tabela
            dados_tabela = []
            anos = sorted(st.session_state.resultados.keys())

            for ano in anos:
                resultado = st.session_state.resultados[ano]
                valor_atual = st.session_state.aliquotas_equivalentes[ano]["valor_atual"]
                diferenca = resultado["imposto_devido"] - valor_atual

                dados_tabela.append({
                    "Ano": ano,
                    "CBS (R$)": resultado["cbs"],
                    "IBS (R$)": resultado["ibs"],
                    "Subtotal IVA (R$)": resultado["imposto_bruto"],
                    "Créditos (R$)": resultado["creditos"],
                    "IVA Devido (R$)": resultado["imposto_devido"],
                    "Impostos Atuais (R$)": resultado["impostos_atuais"]["total"],
                    "Total (R$)": resultado["total_devido"],
                    "Alíquota Efetiva (%)": resultado["aliquota_efetiva"] * 100,
                    "Variação (R$)": diferenca
                })

            # Converter para DataFrame
            df_resultados = pd.DataFrame(dados_tabela)

            # Formatar valores
            cols_dinheiro = ["CBS (R$)", "IBS (R$)", "Subtotal IVA (R$)", "Créditos (R$)",
                            "IVA Devido (R$)", "Impostos Atuais (R$)", "Total (R$)", "Variação (R$)"]

            for col in cols_dinheiro:
                df_resultados[col] = df_resultados[col].apply(lambda x: f"R$ {formatar_br(x)}")

            df_resultados["Alíquota Efetiva (%)"] = df_resultados["Alíquota Efetiva (%)"].apply(lambda x: f"{formatar_br(x)}%")

            # Exibir tabela
            st.dataframe(df_resultados.set_index("Ano"), use_container_width=True)

            # Exibir gráficos
            st.subheader("Análise Gráfica")

            # Tabs para organizar os gráficos
            tab_comp, tab_aliq, tab_trans, tab_inc = st.tabs(["Composição Tributária", "Alíquota Efetiva", "Evolução na Transição", "Impacto dos Incentivos"])

            with tab_comp:
                grafico_comp = criar_grafico_comparativo(st.session_state.resultados, "Comparativo de Impostos por Ano")
                if grafico_comp:
                    st.plotly_chart(grafico_comp, use_container_width=True)

            with tab_aliq:
                grafico_aliq = criar_grafico_aliquotas(st.session_state.resultados, "Evolução da Alíquota Efetiva")
                if grafico_aliq:
                    st.plotly_chart(grafico_aliq, use_container_width=True)

            with tab_trans:
                grafico_trans = criar_grafico_transicao(st.session_state.resultados, "Evolução Tributária na Transição")
                if grafico_trans:
                    st.plotly_chart(grafico_trans, use_container_width=True)

            with tab_inc:
                grafico_inc = criar_grafico_incentivos(st.session_state.resultados, "Impacto dos Incentivos Fiscais no ICMS")
                if grafico_inc:
                    st.plotly_chart(grafico_inc, use_container_width=True)

            # Botões para exportar resultados
            st.subheader("Exportar Resultados")
            col_exp1, col_exp2 = st.columns(2)

            with col_exp1:
                if st.button("Exportar para Excel", key="bt_excel"):
                    with st.spinner("Gerando arquivo Excel..."):
                        b64_excel = exportar_excel()
                        if b64_excel:
                            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_excel}" download="simulacao_iva_dual.xlsx">Download do arquivo Excel</a>'
                            st.markdown(href, unsafe_allow_html=True)

            with col_exp2:
                if st.button("Exportar para PDF", key="bt_pdf"):
                    with st.spinner("Gerando arquivo PDF..."):
                        b64_pdf = exportar_pdf()
                        if b64_pdf:
                            href = f'<a href="data:application/pdf;base64,{b64_pdf}" download="simulacao_iva_dual.pdf">Download do arquivo PDF</a>'
                            st.markdown(href, unsafe_allow_html=True)

# Tab de Configurações
elif opcao_sidebar == "Configurações":
    st.title("Configurações do Simulador")

    # Subtabs para diferentes categorias de configuração
    tab1, tab2, tab3, tab4 = st.tabs(["Alíquotas Base", "Fases de Transição", "Setores", "Salvar/Carregar"])

    # Tab Alíquotas Base
    with tab1:
        st.subheader("Alíquotas Base do IVA Dual")
        st.info("Configure as alíquotas básicas do CBS e IBS conforme a LC 214/2025.")

        col_a1, col_a2 = st.columns(2)

        with col_a1:
            cbs_aliq = st.number_input(
                "Alíquota CBS (%)",
                min_value=0.0,
                max_value=20.0,
                value=float(st.session_state.config.aliquotas_base["CBS"] * 100),
                step=0.1,
                format="%.2f",
                key="aliq_cbs"
            )

        with col_a2:
            ibs_aliq = st.number_input(
                "Alíquota IBS (%)",
                min_value=0.0,
                max_value=30.0,
                value=float(st.session_state.config.aliquotas_base["IBS"] * 100),
                step=0.1,
                format="%.2f",
                key="aliq_ibs"
            )

        if st.button("Atualizar Alíquotas Base", key="update_aliq"):
            st.session_state.config.aliquotas_base["CBS"] = cbs_aliq / 100
            st.session_state.config.aliquotas_base["IBS"] = ibs_aliq / 100
            st.success("Alíquotas base atualizadas com sucesso!")

    # Tab Fases de Transição
    with tab2:
        st.subheader("Fases de Transição do IVA Dual")
        st.info("Configure o percentual progressivo de implementação para cada ano (2026-2033).")

        # Criar uma lista para armazenar os valores atualizados
        valores_fases = {}

        # Criar colunas para exibir os inputs
        col1, col2, col3, col4 = st.columns(4)

        cols = [col1, col2, col3, col4]
        anos = sorted(list(st.session_state.config.fase_transicao.keys()))

        # Distribuir os anos entre as colunas
        for i, ano in enumerate(anos):
            col = cols[i % 4]
            with col:
                valor = st.number_input(
                    f"Ano {ano} (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=float(st.session_state.config.fase_transicao[ano] * 100),
                    step=1.0,
                    format="%.1f",
                    key=f"fase_{ano}"
                )
                valores_fases[ano] = valor / 100

        if st.button("Atualizar Fases de Transição", key="update_fases"):
            for ano, valor in valores_fases.items():
                st.session_state.config.fase_transicao[ano] = valor
            st.success("Fases de transição atualizadas com sucesso!")

    # Tab Setores
    with tab3:
        st.subheader("Configurações por Setor")
        st.info("Configure as alíquotas específicas e reduções para cada setor.")

        # Criar uma tabela para edição
        setores_df = pd.DataFrame({
            "Setor": list(st.session_state.config.setores_especiais.keys()),
            "IBS (%)": [setor["IBS"] * 100 for setor in st.session_state.config.setores_especiais.values()],
            "Redução CBS (%)": [setor["reducao_CBS"] * 100 for setor in st.session_state.config.setores_especiais.values()]
        })

        edited_df = st.data_editor(
            setores_df,
            column_config={
                "Setor": st.column_config.TextColumn("Setor", disabled=True),
                "IBS (%)": st.column_config.NumberColumn("IBS (%)", min_value=0.0, max_value=30.0, step=0.1, format="%.2f"),
                "Redução CBS (%)": st.column_config.NumberColumn("Redução CBS (%)", min_value=0.0, max_value=100.0, step=1.0, format="%.1f")
            },
            use_container_width=True,
            hide_index=True
        )

        if st.button("Atualizar Configurações Setoriais", key="update_setores"):
            # Atualizar as configurações com os valores editados
            for i, setor in enumerate(edited_df["Setor"]):
                st.session_state.config.setores_especiais[setor]["IBS"] = edited_df.iloc[i]["IBS (%)"] / 100
                st.session_state.config.setores_especiais[setor]["reducao_CBS"] = edited_df.iloc[i]["Redução CBS (%)"] / 100

            st.success("Configurações setoriais atualizadas com sucesso!")

    # Tab Salvar/Carregar
    with tab4:
        st.subheader("Salvar/Carregar Configurações")
        st.info("Salve suas configurações em um arquivo JSON ou carregue configurações salvas anteriormente.")

        col_s1, col_s2 = st.columns(2)

        with col_s1:
            # Salvar configurações
            nome_arquivo = st.text_input("Nome do arquivo para salvar", "config_simulador.json")

            if st.button("Salvar Configurações", key="salvar_config"):
                try:
                    sucesso = st.session_state.config.salvar_configuracoes(nome_arquivo)
                    if sucesso:
                        st.success(f"Configurações salvas com sucesso em '{nome_arquivo}'!")
                    else:
                        st.error("Não foi possível salvar as configurações.")
                except Exception as e:
                    st.error(f"Erro ao salvar configurações: {str(e)}")

        with col_s2:
            # Carregar configurações
            uploaded_file = st.file_uploader("Carregar arquivo de configurações", type=["json"])

            if uploaded_file is not None:
                try:
                    # Salvar o arquivo temporariamente
                    with open("temp_config.json", "wb") as f:
                        f.write(uploaded_file.getbuffer())

                    # Carregar configurações
                    sucesso = st.session_state.config.carregar_configuracoes("temp_config.json")

                    if sucesso:
                        # Atualizar a calculadora com as novas configurações
                        st.session_state.calculadora_iva = CalculadoraIVADual(st.session_state.config)
                        st.success("Configurações carregadas com sucesso!")
                    else:
                        st.error("Não foi possível carregar as configurações.")
                except Exception as e:
                    st.error(f"Erro ao carregar configurações: {str(e)}")

# Tab Memória de Cálculo
elif opcao_sidebar == "Memória de Cálculo":
    st.title("Memória de Cálculo")

    # Verificar se existem resultados
    if not st.session_state.resultados:
        st.warning("Execute uma simulação primeiro para visualizar a memória de cálculo.")
    else:
        # Selecionar ano para exibir memória de cálculo
        anos = sorted(list(st.session_state.resultados.keys()))
        ano_selecionado = st.selectbox("Selecione o ano para visualizar a memória de cálculo", anos)

        # Obter memória de cálculo
        memoria = st.session_state.memoria_calculo

        if memoria:
            # Exibir em seções expansíveis
            st.subheader(f"Memória de Cálculo - Ano {ano_selecionado}")

            # Validação de dados
            with st.expander("Validação de Dados", expanded=False):
                for linha in memoria.get("validacao", []):
                    st.write(linha)

            # Base tributável
            with st.expander("Base Tributável", expanded=False):
                for linha in memoria.get("base_tributavel", []):
                    st.write(linha)

            # Alíquotas
            with st.expander("Alíquotas", expanded=False):
                for linha in memoria.get("aliquotas", []):
                    st.write(linha)

            # CBS
            with st.expander("Cálculo da CBS", expanded=False):
                for linha in memoria.get("cbs", []):
                    st.write(linha)

            # IBS
            with st.expander("Cálculo do IBS", expanded=False):
                for linha in memoria.get("ibs", []):
                    st.write(linha)

            # Créditos
            with st.expander("Cálculo dos Créditos", expanded=False):
                for linha in memoria.get("creditos", []):
                    st.write(linha)

            # Imposto devido
            with st.expander("Cálculo do Imposto Devido", expanded=False):
                for linha in memoria.get("imposto_devido", []):
                    st.write(linha)

            # Impostos Atuais
            with st.expander("Cálculo dos Impostos Atuais", expanded=True):
                # PIS
                st.markdown("**PIS:**")
                for linha in memoria.get("impostos_atuais", {}).get("PIS", []):
                    st.write(linha)

                # COFINS
                st.markdown("**COFINS:**")
                for linha in memoria.get("impostos_atuais", {}).get("COFINS", []):
                    st.write(linha)

                # ICMS
                st.markdown("**ICMS:**")
                for linha in memoria.get("impostos_atuais", {}).get("ICMS", []):
                    st.write(linha)

                # ISS
                st.markdown("**ISS:**")
                for linha in memoria.get("impostos_atuais", {}).get("ISS", []):
                    st.write(linha)

                # IPI
                st.markdown("**IPI:**")
                for linha in memoria.get("impostos_atuais", {}).get("IPI", []):
                    st.write(linha)

                # Total Impostos Atuais
                st.markdown("**Total Impostos Atuais:**")
                for linha in memoria.get("impostos_atuais", {}).get("total", []):
                    st.write(linha)

            # Créditos Cruzados
            if memoria.get("creditos_cruzados"):
                with st.expander("Créditos Cruzados", expanded=False):
                    for linha in memoria.get("creditos_cruzados", []):
                        st.write(linha)

            # Total Devido
            with st.expander("Total Devido", expanded=False):
                for linha in memoria.get("total_devido", []):
                    st.write(linha)

            # Opção para exportar a memória de cálculo
            if st.button("Exportar Memória de Cálculo", key="export_memoria"):
                try:
                    # Criar uma string formatada com a memória de cálculo
                    texto_memoria = f"MEMÓRIA DE CÁLCULO - ANO {ano_selecionado}\n\n"

                    # Adicionar seções
                    texto_memoria += "=== VALIDAÇÃO DE DADOS ===\n"
                    for linha in memoria.get("validacao", []):
                        texto_memoria += f"{linha}\n"
                    texto_memoria += "\n"

                    texto_memoria += "=== BASE TRIBUTÁVEL ===\n"
                    for linha in memoria.get("base_tributavel", []):
                        texto_memoria += f"{linha}\n"
                    texto_memoria += "\n"

                    texto_memoria += "=== ALÍQUOTAS ===\n"
                    for linha in memoria.get("aliquotas", []):
                        texto_memoria += f"{linha}\n"
                    texto_memoria += "\n"

                    texto_memoria += "=== CÁLCULO DA CBS ===\n"
                    for linha in memoria.get("cbs", []):
                        texto_memoria += f"{linha}\n"
                    texto_memoria += "\n"

                    texto_memoria += "=== CÁLCULO DO IBS ===\n"
                    for linha in memoria.get("ibs", []):
                        texto_memoria += f"{linha}\n"
                    texto_memoria += "\n"

                    texto_memoria += "=== CÁLCULO DOS CRÉDITOS ===\n"
                    for linha in memoria.get("creditos", []):
                        texto_memoria += f"{linha}\n"
                    texto_memoria += "\n"

                    texto_memoria += "=== CÁLCULO DO IMPOSTO DEVIDO ===\n"
                    for linha in memoria.get("imposto_devido", []):
                        texto_memoria += f"{linha}\n"
                    texto_memoria += "\n"

                    texto_memoria += "=== CÁLCULO DOS IMPOSTOS ATUAIS ===\n"

                    texto_memoria += "\n--- PIS ---\n"
                    for linha in memoria.get("impostos_atuais", {}).get("PIS", []):
                        texto_memoria += f"{linha}\n"

                    texto_memoria += "\n--- COFINS ---\n"
                    for linha in memoria.get("impostos_atuais", {}).get("COFINS", []):
                        texto_memoria += f"{linha}\n"

                    texto_memoria += "\n--- ICMS ---\n"
                    for linha in memoria.get("impostos_atuais", {}).get("ICMS", []):
                        texto_memoria += f"{linha}\n"

                    texto_memoria += "\n--- ISS ---\n"
                    for linha in memoria.get("impostos_atuais", {}).get("ISS", []):
                        texto_memoria += f"{linha}\n"

                    texto_memoria += "\n--- IPI ---\n"
                    for linha in memoria.get("impostos_atuais", {}).get("IPI", []):
                        texto_memoria += f"{linha}\n"

                    texto_memoria += "\n--- TOTAL IMPOSTOS ATUAIS ---\n"
                    for linha in memoria.get("impostos_atuais", {}).get("total", []):
                        texto_memoria += f"{linha}\n"
                    texto_memoria += "\n"

                    if memoria.get("creditos_cruzados"):
                        texto_memoria += "=== CRÉDITOS CRUZADOS ===\n"
                        for linha in memoria.get("creditos_cruzados", []):
                            texto_memoria += f"{linha}\n"
                        texto_memoria += "\n"

                    texto_memoria += "=== TOTAL DEVIDO ===\n"
                    for linha in memoria.get("total_devido", []):
                        texto_memoria += f"{linha}\n"

                    # Gerar arquivo para download
                    b64 = base64.b64encode(texto_memoria.encode()).decode()
                    href = f'<a href="data:file/txt;base64,{b64}" download="memoria_calculo_{ano_selecionado}.txt">Baixar arquivo de memória de cálculo</a>'
                    st.markdown(href, unsafe_allow_html=True)

                except Exception as e:
                    st.error(f"Erro ao exportar memória de cálculo: {str(e)}")

# Tab Sobre
elif opcao_sidebar == "Sobre":
    st.title("Sobre o Simulador de Reforma Tributária")

    # Logo e título
    col_logo, col_titulo = st.columns([1, 3])

    with col_logo:
        st.image("https://www.gov.br/receitafederal/pt-br/assuntos/noticias/2022/dezembro/reforma-tributaria/art-reforma-jpg/@@images/image.jpeg", width=150)

    with col_titulo:
        st.markdown("""
        # Simulador do IVA Dual (CBS/IBS)
        ### Conforme a Lei Complementar 214/2025
        
        **Versão:** 1.0.0
        **Data:** 12/04/2025
        """)

    # Descrição
    st.markdown("""
    ## Sobre este Simulador
    
    Este simulador foi desenvolvido para auxiliar empresas e profissionais a compreender e planejar a transição para o novo sistema tributário brasileiro, baseado no IVA Dual (CBS/IBS), conforme estabelecido pela Lei Complementar 214/2025.
    
    O aplicativo permite simular a carga tributária durante o período de transição (2026-2033), considerando as particularidades setoriais, regimes tributários e incentivos fiscais.
    
    ## Principais Características
    
    - **Simulação Completa**: Cálculo detalhado do IVA Dual (CBS e IBS) e dos impostos do sistema atual durante a transição.
    - **Análise Setorial**: Consideração das alíquotas diferenciadas por setor de atividade.
    - **Incentivos Fiscais**: Simulação do impacto de diversos tipos de incentivos fiscais no ICMS.
    - **Créditos Cruzados**: Implementação das regras de aproveitamento de créditos entre o novo e o antigo sistema.
    - **Memória de Cálculo**: Detalhamento passo a passo de todos os cálculos realizados.
    - **Visualização Gráfica**: Gráficos interativos para melhor compreensão dos resultados.
    
    ## Base Legal
    
    - **Emenda Constitucional nº 132/2023**
    - **Lei Complementar nº 214/2025**
    - **Decretos Regulamentadores**
    - **Portarias e Instruções Normativas**
    
    ## Como Utilizar
    
    1. **Dados da Empresa**: Preencha as informações básicas da empresa e configure os parâmetros de ICMS.
    2. **Incentivos Fiscais**: Configure os incentivos fiscais aplicáveis, se houver.
    3. **Simulação**: Execute a simulação para o período desejado.
    4. **Análise**: Consulte os resultados na tabela e nos gráficos gerados.
    5. **Memória de Cálculo**: Verifique o detalhamento dos cálculos na aba específica.
    
    ## Limitações e Observações
    
    - Este simulador é uma ferramenta de apoio e seus resultados devem ser interpretados por profissionais qualificados.
    - As implementações seguem a legislação vigente até a data de desenvolvimento, podendo haver alterações posteriores.
    - Particularidades específicas de determinados setores ou operações podem requerer análises complementares.
    
    ## Suporte e Contato
    
    Para dúvidas, sugestões ou reportar problemas, entre em contato através do e-mail: suporte@expertzy.com.br
    
    ---
    
    © 2025 Expertzy Inteligência Tributária. Todos os direitos reservados.
    """)

    # Informações técnicas
    with st.expander("Informações Técnicas", expanded=False):
        st.markdown("""
        ### Detalhes Técnicos
        
        **Tecnologias Utilizadas:**
        - Python 3.9+
        - Streamlit 1.32.0+
        - Pandas, NumPy, Matplotlib
        - Plotly
        
        **Estrutura do Código:**
        - `app.py`: Aplicação principal (interface Streamlit)
        - `config.py`: Configurações tributárias
        - `calculadoras.py`: Classes de cálculo (CalculadoraTributosAtuais, CalculadoraIVADual)
        - `utils.py`: Funções utilitárias
        
        **Licença de Uso:**
        Este software é proprietário e sua distribuição, modificação ou uso comercial sem autorização é proibido.
        """)

# Rodapé
st.markdown("---")
st.markdown("© 2025 Expertzy Inteligência Tributária. Todos os direitos reservados.")